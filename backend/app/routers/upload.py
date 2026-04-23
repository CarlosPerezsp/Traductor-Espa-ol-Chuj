from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl
import io

from app.database import get_db

router = APIRouter(prefix="/api/upload", tags=["upload"])


@router.post("/excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    # Detect header row and column positions
    headers = {}
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() in ("arabigo", "arábigo"):
                headers["arabigo"] = cell.column
                header_row_idx = cell.row
            if cell.value and str(cell.value).strip().lower() == "chuj":
                headers["chuj"] = cell.column
                header_row_idx = cell.row
        if headers:
            break

    if "arabigo" not in headers or "chuj" not in headers:
        raise HTTPException(
            status_code=422,
            detail="El archivo debe tener columnas con encabezado 'Arabigo' y 'Chuj'",
        )

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        arabigo_val = row[headers["arabigo"] - 1].value
        chuj_val = row[headers["chuj"] - 1].value

        if arabigo_val is None or chuj_val is None:
            continue

        source_text = str(arabigo_val).strip()
        target_text = str(chuj_val).strip()

        if not source_text or not target_text:
            continue

        # Call stored procedure
        result = db.execute(
            text("CALL sp_insert_translation(:src, :tgt, 1, 2, 2, @res)"),
            {"src": source_text, "tgt": target_text},
        )
        result.close()

        row_result = db.execute(text("SELECT @res AS res")).fetchone()
        if row_result and row_result.res == 1:
            inserted += 1
        else:
            skipped += 1

    # Register upload in uploads table
    db.execute(
        text(
            "INSERT INTO uploads (filename, category_id, inserted, skipped) "
            "VALUES (:filename, 2, :inserted, :skipped)"
        ),
        {"filename": file.filename, "inserted": inserted, "skipped": skipped},
    )
    db.commit()

    return {"inserted": inserted, "skipped": skipped, "filename": file.filename}


@router.get("/history")
def upload_history(db: Session = Depends(get_db)):
    rows = db.execute(
        text(
            "SELECT id, filename, category_id, inserted, skipped, uploaded_at "
            "FROM uploads ORDER BY uploaded_at DESC LIMIT 20"
        )
    ).fetchall()
    return [
        {
            "id": r.id,
            "filename": r.filename,
            "category_id": r.category_id,
            "inserted": r.inserted,
            "skipped": r.skipped,
            "uploaded_at": str(r.uploaded_at),
        }
        for r in rows
    ]
